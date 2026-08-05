"""Excel workbook exporter — all required sheets."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import OUTPUT_WORKBOOK

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _style_sheet(ws) -> None:
    if ws.max_row == 0:
        return
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[letter].width = max(12, max_len + 2)


def export_workbook(sheets: dict[str, pd.DataFrame], path: Path | None = None) -> Path:
    """Write multi-sheet Excel workbook with formatting."""
    out = Path(path) if path else OUTPUT_WORKBOOK
    out.parent.mkdir(parents=True, exist_ok=True)

    # Ensure required sheet order
    ordered = [
        "Data_Ingestion_Log",
        "Cleaned_Data_Summary",
        "Anomalies",
        "Forecasts",
        "Clusters",
        "Risk_Scores",
        "Insights_Narrative",
        "Recommendations",
        "Visualizations_Data",
    ]

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for name in ordered:
            df = sheets.get(name, pd.DataFrame({"note": ["No data"]}))
            if df is None or df.empty:
                df = pd.DataFrame({"note": ["No records for this sheet"]})
            # Excel sheet name limit 31
            safe = name[:31]
            df.to_excel(writer, sheet_name=safe, index=False)
            _style_sheet(writer.sheets[safe])

        # Optional extras
        for name, df in sheets.items():
            if name not in ordered and df is not None and not df.empty:
                safe = name[:31]
                df.to_excel(writer, sheet_name=safe, index=False)
                _style_sheet(writer.sheets[safe])

    logger.info("Workbook written: %s", out)
    return out
