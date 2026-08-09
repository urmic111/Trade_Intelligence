"""
Excel dashboard builder — reads pipeline CSV exports and produces a
formatted multi-sheet dashboard workbook with KPI tiles and embedded charts.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config.settings import OUTPUTS

logger = logging.getLogger(__name__)

CSV_DIR = OUTPUTS / "csv"
DASHBOARD_PATH = OUTPUTS / "Trade_Flow_Dashboard.xlsx"

# Visual theme
NAVY = "0B1F3A"
TEAL = "0E7C7B"
AMBER = "C47B0A"
CRIMSON = "A11D33"
SLATE = "4A5568"
LIGHT = "F2F5F8"
WHITE = "FFFFFF"
CARD = "E8EEF4"

THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _load_csvs(csv_dir: Path) -> dict[str, pd.DataFrame]:
    """Load all sheet CSVs (skip manifest). Key = stem without extension."""
    data: dict[str, pd.DataFrame] = {}
    if not csv_dir.exists():
        raise FileNotFoundError(f"CSV directory not found: {csv_dir}. Run excel_to_csv first.")
    for path in sorted(csv_dir.glob("*.csv")):
        if path.name.startswith("_"):
            continue
        data[path.stem] = pd.read_csv(path)
        logger.info("Loaded %s (%s rows)", path.name, len(data[path.stem]))
    return data


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 11, color: str = NAVY, name: str = "Calibri") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _kpi_card(ws, cell: str, title: str, value: str, accent: str) -> None:
    """Write a 2-row KPI block starting at `cell` spanning 2 columns."""
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    col_letter, row = coordinate_from_string(cell)
    col = column_index_from_string(col_letter)
    title_cell = ws.cell(row=row, column=col, value=title)
    value_cell = ws.cell(row=row + 1, column=col, value=value)
    title_cell.font = _font(bold=True, size=9, color=SLATE)
    value_cell.font = _font(bold=True, size=18, color=accent)
    title_cell.fill = _fill(CARD)
    value_cell.fill = _fill(CARD)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell.alignment = Alignment(horizontal="left", vertical="center")
    # Merge across 2 cols for breathing room
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    for r in (row, row + 1):
        for c in (col, col + 1):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).fill = _fill(CARD)


def _style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN


def _write_table(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    """Write dataframe as a styled table. Returns next empty row."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if r_idx == 0:
                cell.fill = _fill(NAVY)
                cell.font = _font(bold=True, size=10, color=WHITE)
            else:
                cell.font = _font(size=10, color=NAVY)
                if r_idx % 2 == 0:
                    cell.fill = _fill(LIGHT)
    return start_row + len(df) + 1


def _build_dashboard_home(wb: Workbook, data: dict[str, pd.DataFrame]) -> None:
    ws = wb.active
    ws.title = "Dashboard"

    # Banner
    ws.merge_cells("A1:L1")
    banner = ws["A1"]
    banner.value = "TRADE FLOW INTELLIGENCE — EXECUTIVE DASHBOARD"
    banner.font = Font(name="Calibri", bold=True, size=20, color=WHITE)
    banner.fill = _fill(NAVY)
    banner.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:L2")
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A2"].value = f"Source: pipeline CSV exports  ·  Generated {ts}  ·  Data-driven Excel dashboard"
    ws["A2"].font = _font(size=10, color=SLATE)
    ws["A2"].fill = _fill(LIGHT)
    ws.row_dimensions[2].height = 20

    risk = data.get("Risk_Scores", pd.DataFrame())
    anomalies = data.get("Anomalies", pd.DataFrame())
    forecasts = data.get("Forecasts", pd.DataFrame())
    ports = data.get("Port_Risk", pd.DataFrame())
    recs = data.get("Recommendations", pd.DataFrame())
    ingestion = data.get("Data_Ingestion_Log", pd.DataFrame())

    n_critical = int((risk["risk_tier"] == "Critical").sum()) if "risk_tier" in risk.columns else 0
    n_elevated = int((risk["risk_tier"] == "Elevated").sum()) if "risk_tier" in risk.columns else 0
    n_anom = len(anomalies)
    n_fc = int(forecasts["hs_code"].nunique()) if "hs_code" in forecasts.columns else 0
    avg_dep = round(float(risk["dependency_score_0_100"].mean()), 1) if "dependency_score_0_100" in risk.columns and len(risk) else 0
    max_port = round(float(ports["port_congestion_score_0_100"].max()), 1) if "port_congestion_score_0_100" in ports.columns and len(ports) else 0
    n_sources = len(ingestion)

    # KPI row
    _kpi_card(ws, "A4", "DATA SOURCES", str(n_sources), TEAL)
    _kpi_card(ws, "C4", "CRITICAL RISKS", str(n_critical), CRIMSON)
    _kpi_card(ws, "E4", "ELEVATED RISKS", str(n_elevated), AMBER)
    _kpi_card(ws, "G4", "ANOMALIES", str(n_anom), AMBER)
    _kpi_card(ws, "I4", "AVG DEPENDENCY", f"{avg_dep}", TEAL)
    _kpi_card(ws, "K4", "MAX PORT SCORE", f"{max_port}", CRIMSON if max_port >= 60 else TEAL)
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 28

    # Subtitle
    ws.merge_cells("A7:L7")
    ws["A7"].value = f"Forecast coverage: top {n_fc} commodities  ·  Recommendations: {len(recs)}"
    ws["A7"].font = _font(bold=True, size=11, color=NAVY)

    # --- Chart data: Top 12 composite risks ---
    ws["A9"] = "TOP COMPOSITE RISK SCORES"
    ws["A9"].font = _font(bold=True, size=12, color=NAVY)
    ws.merge_cells("A9:D9")

    risk_top = risk.nlargest(12, "composite_risk_0_100")[
        ["commodity", "composite_risk_0_100", "top_supplier", "risk_tier"]
    ].copy() if len(risk) and "composite_risk_0_100" in risk.columns else pd.DataFrame()

    chart_start = 10
    if not risk_top.empty:
        # Short labels for chart axis
        risk_top = risk_top.copy()
        risk_top["label"] = risk_top["commodity"].astype(str).str.slice(0, 28)
        display = risk_top[["label", "composite_risk_0_100"]].rename(
            columns={"label": "Commodity", "composite_risk_0_100": "Risk Score"}
        )
        _write_table(ws, display, chart_start, 1)

        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Composite Supply-Chain Risk (0–100)"
        chart.y_axis.title = None
        chart.x_axis.title = "Risk Score"
        data_ref = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_start + len(display))
        cats = Reference(ws, min_col=1, min_row=chart_start + 1, max_row=chart_start + len(display))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 18
        chart.height = 10
        chart.legend = None
        ws.add_chart(chart, "E10")

    # --- Port congestion table + chart ---
    ws["A26"] = "PORT CONGESTION INDICATORS"
    ws["A26"].font = _font(bold=True, size=12, color=NAVY)
    ws.merge_cells("A26:D26")

    if len(ports) and "port_congestion_score_0_100" in ports.columns:
        port_tbl = ports.sort_values("port_congestion_score_0_100", ascending=False)[
            ["port", "port_congestion_score_0_100", "avg_dwell_days", "avg_congestion"]
        ].copy()
        port_tbl.columns = ["Port", "Congestion Score", "Avg Dwell Days", "Avg Congestion Idx"]
        _write_table(ws, port_tbl, 27, 1)

        pchart = BarChart()
        pchart.type = "col"
        pchart.style = 10
        pchart.title = "Port Congestion Score"
        data_ref = Reference(ws, min_col=2, min_row=27, max_row=27 + len(port_tbl))
        cats = Reference(ws, min_col=1, min_row=28, max_row=27 + len(port_tbl))
        pchart.add_data(data_ref, titles_from_data=True)
        pchart.set_categories(cats)
        pchart.width = 15
        pchart.height = 8
        pchart.legend = None
        ws.add_chart(pchart, "F27")

    # --- Top recommendations snippet ---
    ws["A42"] = "PRIORITY RECOMMENDATIONS"
    ws["A42"].font = _font(bold=True, size=12, color=NAVY)
    ws.merge_cells("A42:L42")

    if len(recs):
        cols = [c for c in ["priority", "audience", "theme", "commodity", "recommendation"] if c in recs.columns]
        _write_table(ws, recs[cols].head(8), 43, 1)

    _set_col_widths(
        ws,
        {
            "A": 28, "B": 16, "C": 14, "D": 18, "E": 14, "F": 14,
            "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14,
        },
    )
    ws.sheet_view.showGridLines = False


def _build_risk_sheet(wb: Workbook, data: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Risk_Detail")
    ws["A1"] = "RISK SCORES — FULL DETAIL"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws.merge_cells("A1:Q1")
    ws.row_dimensions[1].height = 28

    risk = data.get("Risk_Scores", pd.DataFrame())
    if len(risk):
        _write_table(ws, risk, 3, 1)
        for i, col in enumerate(risk.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(22, max(12, len(str(col)) + 2))

        # Dependency share chart data
        if "top_supplier_share_pct" in risk.columns:
            top = risk.nlargest(10, "dependency_score_0_100")[
                ["commodity", "top_supplier_share_pct"]
            ].copy()
            top["commodity"] = top["commodity"].astype(str).str.slice(0, 28)
            start = 3 + len(risk) + 3
            ws.cell(row=start - 1, column=1, value="TOP SUPPLIER SHARE %").font = _font(bold=True, size=12)
            _write_table(ws, top.rename(columns={"commodity": "Commodity", "top_supplier_share_pct": "Top Supplier %"}), start, 1)
            chart = BarChart()
            chart.type = "bar"
            chart.title = "Top Supplier Concentration (%)"
            chart.style = 12
            data_ref = Reference(ws, min_col=2, min_row=start, max_row=start + len(top))
            cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(top))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 16
            chart.height = 9
            chart.legend = None
            ws.add_chart(chart, "E" + str(start))


def _build_forecast_sheet(wb: Workbook, data: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Forecast_Trends")
    ws["A1"] = "IMPORT FORECAST TRENDS — TOP COMMODITIES"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28

    fc = data.get("Forecasts", pd.DataFrame())
    viz = data.get("Visualizations_Data", pd.DataFrame())

    # Prefer import ARIMA forecasts, pivot top 5 commodities for a line chart
    if len(fc) and "model" in fc.columns:
        imp = fc[fc["model"] == "ARIMA(1,1,1)"].copy()
        if "forecast_date" in imp.columns:
            imp["forecast_date"] = pd.to_datetime(imp["forecast_date"], errors="coerce")
        top_hs = (
            imp.groupby("hs_code")["historical_12m_import_usd_m"].first().nlargest(5).index
            if "historical_12m_import_usd_m" in imp.columns
            else imp["hs_code"].unique()[:5]
        )
        subset = imp[imp["hs_code"].isin(top_hs)].copy()
        if len(subset):
            pivot = subset.pivot_table(
                index="forecast_date",
                columns="commodity",
                values="forecast_import_usd_m",
                aggfunc="first",
            ).sort_index()
            pivot.index = pivot.index.astype(str)
            pivot = pivot.reset_index().rename(columns={"forecast_date": "Date"})
            # Truncate column names
            pivot.columns = [str(c)[:32] for c in pivot.columns]
            _write_table(ws, pivot, 3, 1)

            chart = LineChart()
            chart.title = "12-Month Import Forecast (USD millions)"
            chart.style = 10
            chart.y_axis.title = "USD millions"
            chart.x_axis.title = "Month"
            chart.width = 18
            chart.height = 10
            data_ref = Reference(ws, min_col=2, max_col=pivot.shape[1], min_row=3, max_row=3 + len(pivot))
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(pivot))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "A" + str(3 + len(pivot) + 2))

            for i in range(1, pivot.shape[1] + 1):
                ws.column_dimensions[get_column_letter(i)].width = 16

    # Historical viz line data sample
    if len(viz) and "chart" in viz.columns:
        hist = viz[viz["chart"] == "line_top10_imports"].copy()
        if len(hist):
            start = 40
            ws.cell(row=start, column=1, value="HISTORICAL IMPORT SERIES (from Visualizations_Data CSV)").font = _font(
                bold=True, size=12
            )
            sample = hist[["series", "x", "y", "hs_code"]].head(200)
            _write_table(ws, sample, start + 1, 1)


def _build_anomalies_sheet(wb: Workbook, data: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Anomalies")
    ws["A1"] = "TRADE FLOW ANOMALIES"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill = _fill(CRIMSON)
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 28

    anomalies = data.get("Anomalies", pd.DataFrame())
    if len(anomalies):
        _write_table(ws, anomalies, 3, 1)
        for i, col in enumerate(anomalies.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(24, max(12, len(str(col)) + 2))

        if "severity_score" in anomalies.columns:
            chart_df = anomalies.nlargest(min(10, len(anomalies)), "severity_score")[
                ["commodity", "severity_score"]
            ].copy()
            chart_df["commodity"] = chart_df["commodity"].astype(str).str.slice(0, 28)
            start = 3 + len(anomalies) + 3
            ws.cell(row=start - 1, column=1, value="SEVERITY RANKING").font = _font(bold=True, size=12)
            _write_table(
                ws,
                chart_df.rename(columns={"commodity": "Commodity", "severity_score": "Severity"}),
                start,
                1,
            )
            chart = BarChart()
            chart.type = "col"
            chart.title = "Anomaly Severity Scores"
            chart.style = 11
            data_ref = Reference(ws, min_col=2, min_row=start, max_row=start + len(chart_df))
            cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(chart_df))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 14
            chart.height = 8
            chart.legend = None
            ws.add_chart(chart, "D" + str(start))


def _build_insights_sheet(wb: Workbook, data: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Insights")
    ws["A1"] = "NARRATIVE INSIGHTS"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill = _fill(TEAL)
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 28

    narrative = data.get("Insights_Narrative", pd.DataFrame())
    if len(narrative):
        _write_table(ws, narrative, 3, 1)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 100
        ws.column_dimensions["C"].width = 22
        for row in ws.iter_rows(min_row=4, max_row=3 + len(narrative), min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row[0].row].height = 60

    recs = data.get("Recommendations", pd.DataFrame())
    if len(recs):
        start = 3 + len(narrative) + 3
        ws.cell(row=start - 1, column=1, value="ACTIONABLE RECOMMENDATIONS").font = _font(bold=True, size=12, color=TEAL)
        _write_table(ws, recs, start, 1)
        for i in range(1, len(recs.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = max(
                ws.column_dimensions[get_column_letter(i)].width or 12, 14
            )


def _build_clusters_sheet(wb: Workbook, data: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Clusters")
    ws["A1"] = "COUNTRY & COMMODITY CLUSTERS"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    clusters = data.get("Clusters", pd.DataFrame())
    if len(clusters):
        _write_table(ws, clusters, 3, 1)
        for i, col in enumerate(clusters.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(28, max(12, len(str(col)) + 2))

        # Cluster size summary for chart
        if "cluster_label" in clusters.columns:
            sizes = (
                clusters.groupby(["entity_type", "cluster_label"], as_index=False)
                .size()
                .rename(columns={"size": "count"})
            )
            sizes["label"] = sizes["entity_type"] + ": " + sizes["cluster_label"].astype(str)
            start = 3 + len(clusters) + 3
            ws.cell(row=start - 1, column=1, value="CLUSTER SIZES").font = _font(bold=True, size=12)
            chart_df = sizes[["label", "count"]].rename(columns={"label": "Cluster", "count": "Members"})
            _write_table(ws, chart_df, start, 1)
            chart = BarChart()
            chart.type = "col"
            chart.title = "Cluster Membership Counts"
            chart.style = 10
            data_ref = Reference(ws, min_col=2, min_row=start, max_row=start + len(chart_df))
            cats = Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(chart_df))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 14
            chart.height = 8
            chart.legend = None
            ws.add_chart(chart, "D" + str(start))


def _build_data_catalog(wb: Workbook, data: dict[str, pd.DataFrame], csv_dir: Path) -> None:
    ws = wb.create_sheet("Data_Catalog")
    ws["A1"] = "CSV DATA CATALOG (dashboard inputs)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    ws["A1"].fill = _fill(NAVY)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    rows = []
    for name, df in sorted(data.items()):
        rows.append(
            {
                "csv_file": f"{name}.csv",
                "path": str(csv_dir / f"{name}.csv"),
                "rows": len(df),
                "columns": len(df.columns),
            }
        )
    catalog = pd.DataFrame(rows)
    _write_table(ws, catalog, 3, 1)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    ingestion = data.get("Data_Ingestion_Log", pd.DataFrame())
    if len(ingestion):
        ws.cell(row=3 + len(catalog) + 2, column=1, value="INGESTION LOG").font = _font(bold=True, size=12)
        _write_table(ws, ingestion, 3 + len(catalog) + 3, 1)


def build_dashboard(
    csv_dir: Path | None = None,
    output_path: Path | None = None,
) -> Path:
    """Load CSVs and write the Excel dashboard workbook."""
    src = Path(csv_dir) if csv_dir else CSV_DIR
    out = Path(output_path) if output_path else DASHBOARD_PATH
    data = _load_csvs(src)

    wb = Workbook()
    _build_dashboard_home(wb, data)
    _build_risk_sheet(wb, data)
    _build_forecast_sheet(wb, data)
    _build_anomalies_sheet(wb, data)
    _build_insights_sheet(wb, data)
    _build_clusters_sheet(wb, data)
    _build_data_catalog(wb, data, src)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info("Dashboard written: %s", out)
    return out


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
    build_dashboard()
